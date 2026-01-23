from __future__ import annotations

from pathlib import Path
from typing import List, Optional

import pandas as pd

from processors.height import HeightResult
from processors.area import AreaResult
from processors.vkf_rules import small_building_comment, storey_area_comment, vertical_escape_routes


def _build_rows(
    height_result: HeightResult,
    area_result: AreaResult,
    extra_columns: Optional[dict[str, str]] = None,
) -> List[dict[str, str]]:
    """Erzeugt die feste Tabellenstruktur mit Überschriften und Antworten."""
    extra_columns = extra_columns or {}

    def answer(label: str) -> str:
        return extra_columns.get(label, "-")

    rows: List[dict[str, str]] = []

    rows.append({"Beschrieb": "Objektinformationen", "Antwort/Wert": "", "VKF": ""})
    rows.append({"Beschrieb": "Nutzung", "Antwort/Wert": answer("Nutzung"), "VKF": ""})
    rows.append(
        {
            "Beschrieb": "Gebäudehöhe",
            "Antwort/Wert": height_result.rounded_height_m if height_result.height_m is not None else "n/a",
            "VKF": height_result.vkf_category,
        }
    )
    if area_result.building_area_m2 is not None:
        total_area_value = area_result.rounded_area_m2
    else:
        total_area_value = "n/a"
    vkf_comment = small_building_comment(area_result.building_area_m2)

    rows.append(
        {
            "Beschrieb": "Geschossfläche",
            "Antwort/Wert": total_area_value,
            "VKF": vkf_comment,
        }
    )

    if area_result.storeys:
        storeys = sorted(
            area_result.storeys,
            key=lambda s: (s.elevation is None, s.elevation if s.elevation is not None else 0.0),
        )
        for storey in storeys:
            label = storey.name or "Geschoss"
            if storey.elevation is not None:
                label += f" (z = {storey.elevation:.2f} m)"
            storey_vkf = storey_area_comment(storey.area_m2)
            rows.append(
                {
                    "Beschrieb": f"  - {label}",
                    "Antwort/Wert": round(storey.area_m2, 3),
                    "VKF": storey_vkf,
                }
            )
    rows.append({"Beschrieb": "Bauweise", "Antwort/Wert": answer("Bauweise"), "VKF": ""})
    rows.append({"Beschrieb": "Sicherheitsabstand", "Antwort/Wert": answer("Sicherheitsabstand"), "VKF": ""})

    rows.append({"Beschrieb": "Qualitätssicherung", "Antwort/Wert": "", "VKF": ""})
    rows.append({"Beschrieb": "QS-Stufe", "Antwort/Wert": answer("QS-Stufe"), "VKF": ""})
    rows.append({"Beschrieb": "Besonderes", "Antwort/Wert": answer("Besonderes"), "VKF": ""})

    rows.append({"Beschrieb": "Gebäudehülle", "Antwort/Wert": "", "VKF": ""})
    rows.append({"Beschrieb": "Tragwerk", "Antwort/Wert": answer("Tragwerk"), "VKF": ""})
    rows.append({"Beschrieb": "Tragwerk Treppenhaus", "Antwort/Wert": answer("Tragwerk Treppenhaus"), "VKF": ""})
    rows.append({"Beschrieb": "Geschossdecke", "Antwort/Wert": answer("Geschossdecke"), "VKF": ""})
    rows.append({"Beschrieb": "horz. Fluchtweg / Wände", "Antwort/Wert": answer("horz. Fluchtweg / Wände"), "VKF": ""})

    return rows


def write_result_to_excel(
    height_result: HeightResult,
    area_result: AreaResult,
    excel_path: str,
    extra_columns: Optional[dict[str, str]] = None,
) -> None:
    excel_path = Path(excel_path)
    rows = _build_rows(height_result, area_result, extra_columns)
    df = pd.DataFrame(rows, columns=["Beschrieb", "Antwort/Wert", "VKF"])
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(excel_path, index=False)
    _apply_header_formatting(excel_path, rows)


def _apply_header_formatting(excel_path: Path, rows: List[dict[str, str]]) -> None:
    """Setzt Überschriften fett (benötigt openpyxl)."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
    except ModuleNotFoundError:
        return

    header_rows = [
        idx
        for idx, row in enumerate(rows, start=2)
        if row["Beschrieb"] and not row["Antwort/Wert"] and not row["VKF"]
    ]
    if not header_rows:
        return

    wb = load_workbook(excel_path)
    ws = wb.active
    bold_font = Font(bold=True)
    for row_idx in header_rows:
        for col in ("A", "B", "C"):
            ws[f"{col}{row_idx}"].font = bold_font
    wb.save(excel_path)


def write_result_to_template(
    height_result: HeightResult,
    area_result: AreaResult,
    template_path: str,
    output_path: str,
    extra_columns: Optional[dict[str, str]] = None,
) -> None:
    """
    Schreibt Werte in eine bestehende Excel-Vorlage.

    Aktuell befüllt:
      - Nutzung      -> B2
      - QS-Stufe     -> G2
      - Gebäudehöhe  -> B3
      - Besonderes   -> G3
      - Gebäude-Kategorie (Höhe) -> D3
      - Geschossflächen (kommagetrennt) -> B4
      - Bauweise    -> B5
    """
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        missing = exc.name or "openpyxl"
        raise ModuleNotFoundError(f"Fehlendes Paket: {missing}") from exc

    tmpl = Path(template_path)
    if not tmpl.exists():
        raise FileNotFoundError(f"Template nicht gefunden: {tmpl}")

    extra_columns = extra_columns or {}
    usage_val = extra_columns.get("Nutzung", "nicht bestimmt")
    qs_val = extra_columns.get("QS-Stufe", "nicht bestimmt")
    special_val = extra_columns.get("Besonderes", "nicht bestimmt")
    construction_val = extra_columns.get("Bauweise", "nicht bestimmt")
    occupancy_raw = (extra_columns.get("Personenbelegung", "nein") or "nein").lower()
    occupancy_txt = "Raum >300 Per." if occupancy_raw == "ja" else "Raum <300 Per."
    compliance_vert_routes = extra_columns.get("Vertikale Fluchtwege", "nicht bestimmt")
    compliance_exit = extra_columns.get("Ausgang ins Freie", "nicht bestimmt")
    compliance_width = extra_columns.get("Abmessungen min. 1.20m", "nicht bestimmt")
    compliance_doors = extra_columns.get("Anzahl Ausgänge / Türbreiten", "nicht bestimmt")
    compliance_room_seq = extra_columns.get("Raumabfolge", "nicht bestimmt")
    height_val = (
        f"{height_result.rounded_height_m} m" if height_result.height_m is not None else "n/a"
    )
    total_area_val = (
        area_result.rounded_area_m2 if area_result.building_area_m2 is not None else "n/a"
    )
    vkf_cat = height_result.vkf_category or "n/a"

    # Geschossflächen je Geschoss als Text
    storey_text = ""
    if area_result.storeys:
        parts = []
        for s in area_result.storeys:
            label = s.name or "Geschoss"
            parts.append(f"{label}: {round(s.area_m2, 3)} m2")
        storey_text = ", ".join(parts)

    # Vorlage laden und beschreiben
    wb = load_workbook(tmpl)
    ws = wb.active
    ws["B2"] = usage_val
    ws["G2"] = qs_val
    ws["B3"] = height_val
    ws["G3"] = special_val
    ws["D3"] = vkf_cat
    ws["B4"] = storey_text or total_area_val
    ws["B5"] = construction_val
    # Fassade: Einzelwerte
    ws["B7"] = "Klassifiziertes System: RF3-cr"
    ws["C7"] = "Aussenwandbekleidung: RF3-cr"
    ws["D7"] = "Wärmedämmschicht / Zwischenschicht: RF3-cr"
    ws["E7"] = "Lichtbänder: RF3"
    # Dach: Hinweis in G7
    ws["G7"] = "Schichtaufbau Variante 1"
    # Tragwerk-Konzept und Anforderungen
    concept = extra_columns.get("Tragwerk", "nicht bestimmt")
    tw_ug = extra_columns.get("Tragwerk UG", "nicht bestimmt")
    tw_eg_og = extra_columns.get("Tragwerk EG & OG", "nicht bestimmt")
    tw_dg = extra_columns.get("Tragwerk DG", "nicht bestimmt")
    floors = extra_columns.get("Geschossdecken", "nicht bestimmt")
    stairs = extra_columns.get("Treppenhaus", "nicht bestimmt")
    escape = extra_columns.get("Horizontale Fluchtwege", "nicht bestimmt")
    ws["B9"] = f"Tragwerk-Konzept: {concept}"
    ws["C9"] = f"UG: {tw_ug}, EG&OG: {tw_eg_og}"
    ws["C10"] = ""
    ws["D9"] = f"oberstes Geschoss: {tw_dg}"
    ws["E9"] = f"Geschossdecke: {floors}"
    ws["F9"] = f"Treppenhaus: {stairs}"
    ws["G9"] = f"horz. Fluchtweg / Wände: {escape}"
    ws["C11"] = vertical_escape_routes(area_result.building_area_m2)
    ws["C12"] = "Vertikale Fluchtwege müssen an einen sicheren Ort im Freien führen. Mehrere vertikale Fluchtwege müssen unabhängig voneinander an einen sicheren Ort im Freien führen."
    ws["C13"] = "Die Mindestbreite von horizontalen Fluchtwegen muss 1.2 m betragen."
    ws["F11"] = occupancy_txt
    ws["G12"] = (
        "<50 Per.: ein Ausgang mit 0.9 m\n"
        "<100 Personen: zwei Ausgänge mit je 0.9 m\n"
        "<200 Personen: drei Ausgänge mit je 0.9 m oder zwei Ausgänge mit 0.9 m und 1.2 m\n"
        ">200 Personen: mehrere Ausgänge mit mindestens je 1.2 m\n"
        "Büro/Gewerbe/Industrie: Ausgänge 0.9 m zulässig."
    )
    ws["G13"] = (
        "Innerhalb der Nutzungseinheit darf der Fluchtweg über maximal einen angrenzenden Raum "
        "(z. B. Kombizonen) zu einem horizontalen oder vertikalen Fluchtweg führen."
    )
    # Einhaltung Fluchtwege
    ws["B11"] = compliance_vert_routes
    ws["B12"] = compliance_exit
    ws["B13"] = compliance_width
    ws["F12"] = compliance_doors
    ws["F13"] = compliance_room_seq

    # Werte fett formatieren
    try:
        from openpyxl.styles import Font
    except Exception:
        Font = None
    if Font:
        value_cells = [
            "B2",
            "G2",
            "B3",
            "G3",
            "D3",
            "B4",
            "B5",
            "B7",
            "C7",
            "D7",
            "E7",
            "G7",
            "B9",
            "C9",
            "D9",
            "E9",
            "F9",
            "G9",
            "C11",
            "C12",
            "C13",
            "F11",
            "G12",
            "G13",
            "B11",
            "B12",
            "B13",
            "F12",
            "F13",
        ]
        for cell in value_cells:
            if cell in ws:
                current_font = ws[cell].font or Font()
                ws[cell].font = current_font.copy(bold=True)

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
